from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side

# Открываем первый файл и получаем нужный лист
wb1 = load_workbook('file1.xlsx')
sheet1 = wb1.active

# Читаем данные из ячейки A1 первого файла
data1 = sheet1['A1'].value

# Закрываем первый файл
wb1.close()


# Открываем второй файл и получаем нужный лист
wb2 = load_workbook('file2.xlsx')
sheet2 = wb2.active

# Читаем данные из ячейки A1 второго файла
data2 = sheet2['A1'].value

# Закрываем воторой файл
wb2.close()


# открываем третьи файл и получаем нужный лист
wb3 = load_workbook('file3.xlsx')
sheet3 = wb3.active

# читаем данныу из ячейки A1 третьего файла
data3 = sheet3['A1'].value

# закрываем третьи файл
wb3.close()

print(data1, data2, data3)


# Создаем матрицу
matrix = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]

# Преобразуем матрицу в одномерный список и сортируем его в порядке убывания
sorted_list = sorted([elem for row in matrix for elem in row], reverse=True)

# Возвращаем отсортированный список обратно в матрицу
sorted_matrix = [sorted_list[i:i+len(matrix[0])] for i in range(0, len(sorted_list), len(matrix[0]))]

# Создаем новый Excel-файл и добавляем лист
wb = Workbook()
sheet = wb.active

# Задаем стиль для ячеек
font = Font(bold=True)
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

# Записываем значения в ячейки
for i in range(len(sorted_matrix)):
    for j in range(len(sorted_matrix[0])):
        cell = sheet.cell(row=i+1, column=j+1)
        cell.value = sorted_matrix[i][j]
        cell.font = font
        cell.border = border

# Сохраняем файл
wb.save('sorted_matrix.xlsx')

